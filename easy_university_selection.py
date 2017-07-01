# -*- coding: utf-8 -*-
# ////////////////////////////////////////////////////////////////////
# //                          _ooOoo_                               //
# //                         o8888888o                              //
# //                         88" . "88                              //
# //                         (| ^_^ |)                              //
# //                         O\  =  /O                              //
# //                      ____/`---'\____                           //
# //                    .'  \\|     |//  `.                         //
# //                   /  \\|||  :  |||//  \                        //
# //                  /  _||||| -:- |||||-  \                       //
# //                  |   | \\\  -  /// |   |                       //
# //                  | \_|  ''\---/''  |   |                       //
# //                  \  .-\__  `-`  ___/-. /                       //
# //                ___`. .'  /--.--\  `. . ___                     //
# //              ."" '<  `.___\_<|>_/___.'  >'"".                  //
# //            | | :  `- \`.;`\ _ /`;.`/ - ` : | |                 //
# //            \  \ `-.   \_ __\ /__ _/   .-` /  /                 //
# //      ========`-.____`-.___\_____/___.-`____.-'========         //
# //                           `=---='                              //
# //      ^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^        //
# //         佛祖保佑       永无BUG     永不修改                      //
# ////////////////////////////////////////////////////////////////////

import sys
import os
import urllib2
import xml.sax
from xml.dom.minidom import parse
import xml.dom.minidom
from openpyxl import Workbook
import pickle
import time


# 分数划线
class ScoreLine:
    def __init__(self):
        pass

    year = ''
    region = ''
    subject = ''
    tier = ''
    score = 0


# 省入取分数
class ProvinceScore:
    def __init__(self):
        pass

    year = 0
    region = ''
    school = ''
    subject = ''
    maxScore = 0
    avgScore = 0
    minScore = 0
    tier = ''
    hope = 0
    hot = 0


# 专业分数线
class MajorScore:
    def __init__(self):
        pass

    year = 0
    region = ''
    school = ''
    subject = ''
    maxScore = 0
    avgScore = 0
    minScore = 0
    tier = ''
    hope = 0
    hot = 0
    major = ''
    majorName = ''


# 高校信息
class UniversityInfo:
    def __init__(self):
        pass

    longitude = ''
    latitude = ''
    name = ''
    region = ''
    regionCode = ''
    level = ''
    hot = 0
    classes = ''
    classRank = 0
    web = ''
    code = ''


# 载入所有年份的分数线
def load_score_line():
    files = os.listdir('./resource/score_line/' + regionCode)
    sls = {}  # year:region:subject:tier score
    for sFile in files:
        f = open('./resource/score_line/' + regionCode + '/' + sFile)
        iter_f = iter(f)  # 创建迭代器
        for line in iter_f:
            line = ''.join(line.split())
            arr = line.split(",")
            if len(arr) < 4: continue
            sl = ScoreLine()
            sl.year = arr[0]
            sl.region = regionCodeDict[arr[1]]
            if arr[2] == '理科':
                sl.subject = '10035'
            elif arr[2] == '文科':
                sl.subject = '10034'
            else:
                continue

            if '一' in arr[3]:
                sl.tier = '10036'
            elif '二' in arr[3]:
                sl.tier = '10037'
            elif '三' in arr[3]:
                sl.tier = '10038'
            elif '专科' or '高职' in arr[3]:
                sl.tier = '10148'
            else:
                continue

            # 专为福建没有三本特殊处理
            if regionCode == '10024':
                if sl.tier == '10148':
                    sl.tier = '10038'

            sl.score = int(arr[4])
            sls[sl.year + ',' + sl.region + ',' + sl.subject + ',' + sl.tier] = sl
    return sls


# 加载高校在各省的专业录取分数线
def load_major_score():
    score_path = './resource/spider_files/major_score_line/' + regionCode + '_' + subject + '.dump'
    if os.path.exists(score_path):
        tmp_file = open(score_path, 'rb')
        d = pickle.load(tmp_file)
        tmp_file.close()
        print '共加载' + str(len(d)) + '条专业分数线数据'
        return d
    paths = os.listdir('./resource/spider_files/major_score_line/' + regionCode + '/')
    pss = {}
    for path in paths:
        files = os.listdir('./resource/spider_files/major_score_line/' + regionCode + '/' + path)
        for sFile in files:
            if not os.path.isdir(sFile):
                if not (subject in sFile): continue
                print sFile
                dom = xml.dom.minidom.parse(
                    './resource/spider_files/major_score_line/' + regionCode + '/' + path + '/' + sFile)
                root = dom.documentElement
                major_elements = root.getElementsByTagName("areapiont")
                for major_element in major_elements:
                    year_node = major_element.getElementsByTagName("year")[0]
                    y = ''
                    if len(year_node.childNodes) > 0: y = year_node.childNodes[0].nodeValue

                    max_score_node = major_element.getElementsByTagName("maxfs")[0]
                    max_score = ''
                    if len(max_score_node.childNodes) > 0: max_score = max_score_node.childNodes[0].nodeValue

                    min_score_node = major_element.getElementsByTagName("minfs")[0]
                    min_score = ''
                    if len(min_score_node.childNodes) > 0: min_score = min_score_node.childNodes[0].nodeValue

                    avg_score_node = major_element.getElementsByTagName("varfs")[0]
                    avg_score = ''
                    if len(avg_score_node.childNodes) > 0: avg_score = avg_score_node.childNodes[0].nodeValue

                    tier_node = major_element.getElementsByTagName("pc")[0]
                    tier = ''
                    if len(tier_node.childNodes) > 0: tier = tier_node.childNodes[0].nodeValue

                    major_name_node = major_element.getElementsByTagName("specialname")[0]
                    major_name = ''
                    if len(major_name_node.childNodes) > 0: major_name = major_name_node.childNodes[0].nodeValue

                    ms = MajorScore()
                    if not ('--' == y or '' == y): ms.year = int(y)
                    if ms.year < int(year) - 3: continue
                    if not ('--' == max_score or '' == max_score): ms.maxScore = int(max_score[0:3])
                    if not ('--' == min_score or '' == min_score): ms.minScore = int(min_score[0:3])
                    if not ('--' == avg_score or '' == avg_score): ms.avgScore = int(avg_score[0:3])
                    tier = tier.encode('utf-8')

                    if '一' in tier:
                        tier_code = '10036'
                    elif '二' in tier:
                        tier_code = '10037'
                    elif '三' in tier:
                        tier_code = '10038'
                    elif '专' in tier:
                        tier_code = '10148'
                    elif '提前' in tier:
                        tier_code = '10149'
                    else:
                        continue

                    major_name = major_name.encode('utf-8')

                    ms.tier = tier_code
                    ms.region = regionCode
                    ms.school = path
                    ms.subject = subject
                    ms.majorName = major_name
                    if majorCodeDict.has_key(major_name):
                        ms.major = majorCodeDict[major_name]

                    # 学校 年份 福建 文科 批次 = 清华大学2016年在福建地区文科第一批次招生分数线
                    key = path + ',' + y + ',' + regionCode + ',' + subject + ',' + tier_code + ',' + ms.major
                    pss[key] = ms
    with open(score_path, 'wb') as pickle_file:
        pickle.dump(pss, pickle_file)
        pickle_file.close()
    return pss


# 加载高校在各省的录取分数线
def load_province_score():
    score_path = './resource/spider_files/province_score_line/' + regionCode + '_' + subject + '.dump'
    if os.path.exists(score_path):
        tmp_file = open(score_path, 'rb')
        d = pickle.load(tmp_file)
        tmp_file.close()
        print '共加载' + str(len(d)) + '条省录取分数线数据'
        return d
    paths = os.listdir('./resource/spider_files/province_score_line/' + regionCode + '/')
    pss = {}
    for path in paths:
        files = os.listdir('./resource/spider_files/province_score_line/' + regionCode + '/' + path)
        for sFile in files:
            if not os.path.isdir(sFile):
                if not (subject in sFile): continue
                print sFile
                dom = xml.dom.minidom.parse(
                    './resource/spider_files/province_score_line/' + regionCode + '/' + path + '/' + sFile)
                root = dom.documentElement
                score_elements = root.getElementsByTagName("score")
                for score_element in score_elements:
                    # print score.nodeName
                    year_node = score_element.getElementsByTagName("year")[0]
                    y = ''
                    if len(year_node.childNodes) > 0: y = year_node.childNodes[0].nodeValue
                    # print (yearNode.childNodes)
                    max_score_node = score_element.getElementsByTagName("maxScore")[0]
                    max_score = ''
                    if len(max_score_node.childNodes) > 0: max_score = max_score_node.childNodes[0].nodeValue
                    min_score_node = score_element.getElementsByTagName("minScore")[0]
                    min_score = ''
                    if len(min_score_node.childNodes) > 0: min_score = min_score_node.childNodes[0].nodeValue
                    avg_score_node = score_element.getElementsByTagName("avgScore")[0]
                    avg_score = ''
                    if len(avg_score_node.childNodes) > 0: avg_score = avg_score_node.childNodes[0].nodeValue
                    tier_node = score_element.getElementsByTagName("rb")[0]
                    tier = ''
                    if len(tier_node.childNodes) > 0: tier = tier_node.childNodes[0].nodeValue
                    ps = ProvinceScore()
                    if not ('--' == y or '' == y): ps.year = int(y)
                    if not ('--' == max_score or '' == max_score): ps.maxScore = int(max_score[0:3])
                    if not ('--' == min_score or '' == min_score): ps.minScore = int(min_score[0:3])
                    if not ('--' == avg_score or '' == avg_score): ps.avgScore = int(avg_score[0:3])
                    tier = tier.encode('utf-8')

                    if '一' in tier:
                        tier_code = '10036'
                    elif '二' in tier:
                        tier_code = '10037'
                    elif '三' in tier:
                        tier_code = '10038'
                    elif '专' in tier:
                        tier_code = '10148'
                    else:
                        continue

                    ps.tier = tier_code
                    ps.region = regionCode
                    ps.school = path
                    ps.subject = subject

                    # 学校 年份 福建 文科 批次 = 清华大学2016年在福建地区文科第一批次招生分数线
                    key = path + ',' + y + ',' + regionCode + ',' + subject + ',' + tier_code
                    pss[key] = ps

    with open(score_path, 'wb') as pickle_file:
        pickle.dump(pss, pickle_file)
        pickle_file.close()
    return pss


def load_university_info():
    dump_file = './resource/university_info.dump'
    university_dict = {}
    if os.path.exists(dump_file):
        f = open(dump_file, 'rb')
        d = pickle.load(f)
        f.close()
        return d

    f = open('./resource/university_info.csv')
    iter_f = iter(f)  # 创建迭代器
    for line in iter_f:
        if line.startswith('#'): continue

        line = ''.join(line.split())
        university = UniversityInfo()
        arr = line.split(',')
        university.longitude = arr[0]
        university.latitude = arr[1]
        university.name = arr[2]
        university.region = arr[3]
        university.regionCode = regionCodeDict[arr[3]]
        university.level = arr[4]
        university.hot = arr[5]
        university.classes = arr[6]
        university.classRank = arr[7]
        university.web = arr[8]
        university.code = arr[9]
        university_dict[arr[9]] = university

    with open(dump_file, 'wb') as pickle_file:
        pickle.dump(university_dict, pickle_file)
        pickle_file.close()
    return university_dict


def init_cvs_kv(path, reverse):
    kv = {}
    f = open(path)
    iter_f = iter(f)  # 创建迭代器
    for line in iter_f:
        line = ''.join(line.split())
        arr = line.split(',')
        if not len(arr) == 2: continue
        if reverse:
            kv[arr[1]] = arr[0]
        else:
            kv[arr[0]] = arr[1]

    return kv


def init_spider(path):
    url_set = set()
    if os.path.exists(path):
        f = open(path)
        iter_f = iter(f)  # 创建迭代器
        for line in iter_f:
            line = ''.join(line.split())
            url_set.add(line)

    return url_set


# 抓取大学专业分
def spider_university_major_score_line(info):
    spider_score_line('./resource/spider_files/major_score_line/',
                      'http://gkcx.eol.cn/commonXML/schoolSpecialPoint/schoolSpecialPoint', 'schoolSpecialPoint', '',info)


# 抓取大学省录取分
def spider_university_province_score_line(tier,info):
    spider_score_line('./resource/spider_files/province_score_line/',
                      'http://gkcx.eol.cn/schoolhtm/scores/provinceScores', 'provinceScores', tier,info)


def spider_score_line(save_path, spider_url, xml_name, tier,info):
    count = 0
    url404 = init_spider(save_path + regionCode + '_404.url')
    has_spider = init_spider(save_path + regionCode + '_spider.url')
    url404_size = len(url404)
    has_spider_size = len(has_spider)
    if '' == tier:
        file_suffix = '.xml'
    else:
        file_suffix = '_' + tier + '.xml'

    url_base = spider_url + '[university_code]_' + regionCode + '_' + subject + file_suffix
    for k in universityInfoDict:
        url = url_base.replace('[university_code]', universityInfoDict[k].code)
        if url in url404: continue
        count = count + 1
        if url in has_spider: continue
        print url
        req = urllib2.Request(url)
        res_data = urllib2.urlopen(req)
        if "http://gkcx.eol.cn/404.htm" == res_data.url:
            url404.add(url)
            continue
        has_spider.add(url)
        res = res_data.read()
        path = save_path + regionCode + '/' + universityInfoDict[k].code
        if not os.path.exists(path): os.makedirs(path)
        xml_file = open(
            path + '/' + xml_name + universityInfoDict[
                k].code + '_' + regionCode + '_' + subject + file_suffix, 'w')
        xml_file.write(res)
        xml_file.close()

    if not len(url404) == url404_size:
        wr = ''
        for U in url404:
            wr = wr + U + '\n'
        tmp_file = open(save_path + regionCode + '_404.url', 'w')
        tmp_file.write(str(wr))  # 写入内容，如果没有该文件就自动创建
        tmp_file.close()  # (关闭文件)

    if not len(has_spider) == has_spider_size:
        wr = ''
        for U in has_spider:
            wr = wr + U + '\n'
        tmp_file = open(save_path + regionCode + '_spider.url', 'w')
        tmp_file.write(str(wr))  # 写入内容，如果没有该文件就自动创建
        tmp_file.close()  # (关闭文件)
    print info + '抓取完成,共抓取' + str(count) + '条数据'


# 筛选高校
def filter_university(scores):
    year_int = int(year)

    last1 = evaluate_score[year_int - 1]
    last2 = evaluate_score[year_int - 2]
    last3 = evaluate_score[year_int - 3]

    result = []
    school_set = set()
    for k in scores:
        s = scores[k]
        is_filter = True
        if s.school in school_set: continue
        if not subject == s.subject: continue
        if s.minScore == 0 and s.avgScore == 0 and s.maxScore == 0: continue
        hope = 0
        if not s.minScore == 0:
            if s.year == year_int - 1:
                if last1 > s.minScore:
                    is_filter = False
                    hope = 3
            elif s.year == year_int - 2:
                if last2 > s.minScore:
                    is_filter = False
                    hope = 2
            elif s.year == year_int - 3:
                if last3 > s.minScore:
                    is_filter = False
                    hope = 1

        if not s.avgScore == 0:
            if s.year == year_int - 1:
                if last1 > s.avgScore:
                    is_filter = False
                    hope = 6
            elif s.year == year_int - 2:
                if last2 > s.avgScore:
                    is_filter = False
                    hope = 5
            elif s.year == year_int - 3:
                if last3 > s.avgScore:
                    is_filter = False
                    hope = 4

        if not s.maxScore == 0:
            if s.year == year_int - 1:
                if last1 > s.maxScore:
                    is_filter = False
                    hope = 9
            elif s.year == year_int - 2:
                if last2 > s.maxScore:
                    is_filter = False
                    hope = 8
            elif s.year == year_int - 3:
                if last3 > s.maxScore:
                    is_filter = False
                    hope = 7

        if is_filter: continue
        school_set.add(s.school + str(s.year))
        s.hope = hope
        s.hot = universityInfoDict[s.school].hot
        result.append(s)
    return result


def filter_university_by_major_score():
    return filter_university(majorScores)


def filter_university_by_province_score():
    return filter_university(provinceScores)


def evaluate_three_year_score():
    # 换算该学生在往年的分数
    # 粗暴的算法
    # 计算 考生的分数在今年高考划线的比值
    # 如 该生文科393 ，2017 划线 489,380,300 ，比值分别是393/489=0.803,393/380=1.034,393/300=1.31
    # 2016 年 划线 501,403,319,评估得分为(501*0.803+403*1.034+319*1.31)/3=412.3,评估该生在2016分数为412.3
    result = {}
    year_int = int(year)
    score1 = scoreLines[str(year_int) + ',' + regionCode + ',' + subject + ',10036'].score
    score2 = scoreLines[str(year_int) + ',' + regionCode + ',' + subject + ',10037'].score
    score3 = scoreLines[str(year_int) + ',' + regionCode + ',' + subject + ',10038'].score
    tier = ''
    if score > score3:
        tier = '10038'
        if score > score2:
            tier = '10037'
            if score > score1:
                tier = '10036'
    # if '' == tier:
    #   print 'error....'
    #   return

    rate1 = score / float(score1)
    rate2 = score / float(score2)
    rate3 = score / float(score3)  # 福建地区没有三本划线，取专科划线

    # 改进算法，分数所在批次权重为0.7,其他为0.15
    # TODO 考虑结合一分一段排名表来评估分数，暂时没有找到数据，待完成
    last1 = evaluate_score(year_int, regionCode, subject, tier, 1, rate1, rate2, rate3)
    result[year_int - 1] = last1
    print '评估' + str(year_int - 1) + '分数为：' + str(last1)

    last2 = evaluate_score(year_int, regionCode, subject, tier, 2, rate1, rate2, rate3)
    result[year_int - 2] = last2
    print '评估' + str(year_int - 2) + '分数为：' + str(last2)

    last3 = evaluate_score(year_int, regionCode, subject, tier, 3, rate1, rate2, rate3)
    result[year_int - 3] = last3
    print '评估' + str(year_int - 3) + '分数为：' + str(last3)
    print '%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%'

    return result


def evaluate_score(year, region, subject, tier, n, rate1, rate2, rate3):
    if '10036' == tier:
        last1score1 = scoreLines[str(year - n) + ',' + region + ',' + subject + ',10036'].score * rate1 * 0.7
    else:
        last1score1 = scoreLines[str(year - n) + ',' + region + ',' + subject + ',10036'].score * rate1 * 0.15
    if '10037' == tier:
        last1score2 = scoreLines[str(year - n) + ',' + region + ',' + subject + ',10037'].score * rate2 * 0.7
    else:
        last1score2 = scoreLines[str(year - n) + ',' + region + ',' + subject + ',10037'].score * rate2 * 0.15
    if '10038' == tier:
        last1score3 = scoreLines[str(year - n) + ',' + region + ',' + subject + ',10038'].score * rate3 * 0.7
    else:
        last1score3 = scoreLines[str(year - n) + ',' + region + ',' + subject + ',10038'].score * rate3 * 0.15

    if '' == tier:
        last1score1 = scoreLines[str(year - n) + ',' + region + ',' + subject + ',10036'].score * rate1 * 0.3
        last1score2 = scoreLines[str(year - n) + ',' + region + ',' + subject + ',10037'].score * rate2 * 0.3
        last1score3 = scoreLines[str(year - n) + ',' + region + ',' + subject + ',10038'].score * rate3 * 0.4

    return last1score1 + last1score2 + last1score3


def save(title, university, xlsx):
    print '筛选结果如下，结果将保存到' + xlsx
    print '-----------------------------------------------------------------------------------------------------'
    print title
    print '-----------------------------------------------------------------------------------------------------'
    # 筛选结果保存到xls
    wb = Workbook()
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    ws = wb.active
    column_names = title.split('\t')
    for i in range(1, len(column_names) + 1):
        ws.cell(row=1, column=i).value = column_names[i - 1]

    row = 2
    for u in university:
        if u.tier in filterTier: continue
        col_content = universityInfoDict[u.school].name + '\t'
        if isinstance(u, MajorScore):
            col_content = col_content + u.majorName + '\t'
        col_content = col_content + universityInfoDict[u.school].region + '\t' + \
                      universityInfoDict[u.school].classes + '\t' + str(
            universityInfoDict[u.school].classRank) + '\t' + str(u.hot) + '\t' + str(u.hope) + '\t' + str(
            u.maxScore) + '\t' + str(u.minScore) + '\t' + str(u.avgScore) + '\t' + customCodeDict[u.tier] + '\t' + str(
            u.year)

        print col_content
        row_values = col_content.split('\t')
        for col in range(1, len(column_names) + 1):
            ws.cell(row=row, column=col).value = row_values[col - 1]
        row = row + 1
        # 保存
    wb.save(filename=xlsx)


def save_xlsx():
    t = str(time.time())
    save('学校\t地区\t类别\t类别排名\t热度排名\t录取成功预测值（1-9）\t 最高分\t最低分\t平均分\t批次\t年份', universityListByProvinceScore,
         './resource/result/result_by_province_score' + t[0:10] + '.xlsx')
    save('学校\t专业\t地区\t类别\t类别排名\t热度排名\t录取成功预测值（1-9）\t 最高分\t最低分\t平均分\t批次\t年份', universityListByMajorScore,
         './resource/result/result_by_major_score' + t[0:10] + '.xlsx')


def init_custom_code():
    code_dict = {'10035': '理科', '10034': '文科', '10036': '一本', '10037': '二本', '10038': '三本', '10148': '专科',
                 '10149': '提前'}
    return code_dict


def help(region):
    print '输入参数：省份（行政区代码） 文理科 考生分数 年份 过滤批次'
    print '省份代码对应如下'
    for k in region:
        print k + '-' + region[k]
    print '文科-10034，理科10035'
    print '10036 一本,10037 二本,10038 三本,10148 专科'

    sys.exit(-1)


# 10035 理科
# 10034 文科
#
# 10036 一本
# 10037 二本
# 10038 三本
# 10148 专科
# 10149 提前

if __name__ == "__main__":
    regionCodeDict = init_cvs_kv('./resource/region_code.csv', False)
    codeRegionDict = init_cvs_kv('./resource/region_code.csv', True)
    codeMajorDict = init_cvs_kv('./resource/major_code.csv', False)
    majorCodeDict = init_cvs_kv('./resource/major_code.csv', True)
    customCodeDict = init_custom_code()

    # print str(len(sys.argv))
    if len(sys.argv) < 5:
        help(regionCodeDict)
    filterTier = ''
    if len(sys.argv) > 5:
        filterTier = sys.argv[5].split(',')

    regionCode = sys.argv[1]  # region_code 行政区代码
    subject = sys.argv[2]  # arts_or_science_code 文理科
    score = int(sys.argv[3])  # 考生分数
    year = sys.argv[4]  # 年份
    # tierCode = sys.argv[3]  # tier_code 本科层次（一本，二本，三本，提前，专科）
    print '%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%'
    print '年份：' + year
    print '地区：' + codeRegionDict[regionCode]
    print '分数：' + sys.argv[3] + ' ' + customCodeDict[subject]
    if not filterTier == '':
        for f in filterTier:
            print '过滤：' + customCodeDict[f]
    print '%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%'

    universityInfoDict = load_university_info()
    print '加载高校库完成，共有' + str(len(universityInfoDict)) + '所高校信息载入'
    print '%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%'
    print '抓取高校库中所有高校在[' + codeRegionDict[regionCode] + ']地区[' + customCodeDict[subject] + ']招生分数线'
    spider_university_province_score_line('10036','本一批次')
    spider_university_province_score_line('10037','本二批次')
    spider_university_province_score_line('10038','本三批次')
    spider_university_province_score_line('10148','高职专科批次')
    print '%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%'
    print '抓取高校库中所有高校在[' + codeRegionDict[regionCode] + ']地区[' + customCodeDict[subject] + ']专业分数线'
    spider_university_major_score_line('')
    print '%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%'
    print '载入[' + codeRegionDict[regionCode] + ']地区历年高考划线'
    scoreLines = load_score_line()  # 历年分数线
    print '载入全国高校在[' + codeRegionDict[regionCode] + ']地区[' + customCodeDict[subject] + ']历年录取分数线'
    provinceScores = load_province_score()  # 各学校入取分数
    print '载入全国高校在[' + codeRegionDict[regionCode] + ']地区[' + customCodeDict[subject] + ']历年录取专业分数线'
    majorScores = load_major_score()  # 各学校各专业录取分数
    print '%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%%'
    # 评估分数
    evaluate_score = evaluate_three_year_score()
    # 筛选高校
    universityListByProvinceScore = filter_university_by_province_score()
    universityListByMajorScore = filter_university_by_major_score()
    # 保存到 xlsx
    save_xlsx()
